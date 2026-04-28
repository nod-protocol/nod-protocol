"""Build taxonomy/naics-2022.json from Census Bureau XLSX sources.

Usage:
    # 1. Download the source files (or pass --download to fetch them)
    curl -sSL -o /tmp/naics-codes.xlsx \\
        'https://www.census.gov/naics/2022NAICS/2-6%20digit_2022_Codes.xlsx'
    curl -sSL -o /tmp/naics-descriptions.xlsx \\
        'https://www.census.gov/naics/2022NAICS/2022_NAICS_Descriptions.xlsx'

    # 2. Run the build (paths and output overridable via env vars)
    NAICS_CODES=/tmp/naics-codes.xlsx \\
    NAICS_DESCS=/tmp/naics-descriptions.xlsx \\
        python3 scripts/build-naics.py

Requires: openpyxl  (pip install openpyxl)
"""
import openpyxl, json, re, os

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CODES_PATH = os.environ.get('NAICS_CODES', '/tmp/naics-src/codes.xlsx')
DESCS_PATH = os.environ.get('NAICS_DESCS', '/tmp/naics-src/descriptions.xlsx')
OUT_PATH = os.environ.get(
    'NAICS_OUT', os.path.join(REPO_ROOT, 'taxonomy', 'naics-2022.json')
)

LEVELS = {
    2: "sector",
    3: "subsector",
    4: "industry-group",
    5: "industry",
    6: "national-industry",
}
# Census represents three sectors as ranges of 2-digit codes.
RANGE_SECTORS = {"31-33", "44-45", "48-49"}
# Map each constituent 2-digit prefix to its range sector.
SECTOR_RANGE_MAP = {
    "31": "31-33", "32": "31-33", "33": "31-33",
    "44": "44-45", "45": "44-45",
    "48": "48-49", "49": "48-49",
}
SEE_RE = re.compile(r'See industry description for (\d+)\.', re.IGNORECASE)


def short_desc(raw):
    """Take the first content paragraph; drop Cross-References tail."""
    if not raw:
        return None
    text = str(raw).strip()
    cr = text.find("Cross-References")
    if cr >= 0:
        text = text[:cr].rstrip(" \n\t.")
    paras = [p.strip() for p in re.split(r'\n\s*\n', text) if p.strip()]
    if not paras:
        return None
    # skip header-like first para ("The Sector as a Whole")
    if len(paras) > 1 and len(paras[0]) < 80 and "." not in paras[0]:
        return paras[1]
    return paras[0]


# --- Load codes (canonical titles) ---
wb = openpyxl.load_workbook(CODES_PATH, read_only=True)
ws = wb.active
codes_rows = []
for row in ws.iter_rows(min_row=3, values_only=True):
    seq, code, title = row[0], row[1], row[2]
    if code is None or title is None:
        continue
    codes_rows.append((str(code), str(title).strip()))
wb.close()

# --- Load descriptions ---
wb = openpyxl.load_workbook(DESCS_PATH, read_only=True)
ws = wb.active
descs_raw = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    code, _title, desc = row[0], row[1], row[2]
    if code is None:
        continue
    descs_raw[str(code)] = desc
wb.close()

# --- Build entries ---
entries = []
odd_lengths = []
for code, title in codes_rows:
    if code in RANGE_SECTORS:
        level = "sector"
        parent = None
    else:
        level = LEVELS.get(len(code))
        if level is None:
            odd_lengths.append((code, title))
            continue
        if len(code) == 2:
            parent = None
        else:
            raw_parent = code[:-1]
            # Subsectors of range sectors: remap '311' -> '31-33' (since '31' isn't a row).
            parent = SECTOR_RANGE_MAP.get(raw_parent, raw_parent)
    desc = short_desc(descs_raw.get(code))
    if desc:
        m = SEE_RE.search(desc)
        if m:
            ref_desc = short_desc(descs_raw.get(m.group(1)))
            if ref_desc:
                desc = ref_desc
    entries.append({
        "code": code,
        "title": title,
        "parent_code": parent,
        "level": level,
        "description": desc,
    })

# --- Sanity checks ---
by_level = {}
for e in entries:
    by_level[e["level"]] = by_level.get(e["level"], 0) + 1

# Verify parents exist (for non-sector entries)
codes_set = {e["code"] for e in entries}
orphans = [e["code"] for e in entries if e["parent_code"] and e["parent_code"] not in codes_set]

# Count missing descriptions
missing_desc = [e["code"] for e in entries if not e["description"]]

print(f"total entries: {len(entries)}")
print(f"by level: {by_level}")
print(f"odd-length codes skipped: {len(odd_lengths)} {odd_lengths[:5]}")
print(f"orphan parents: {len(orphans)} {orphans[:5]}")
print(f"missing descriptions: {len(missing_desc)} {missing_desc[:5]}")

# --- Write output ---
os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
payload = {
    "version": "NAICS 2022",
    "source": "US Census Bureau",
    "source_url": "https://www.census.gov/naics/",
    "source_files": [
        "https://www.census.gov/naics/2022NAICS/2-6%20digit_2022_Codes.xlsx",
        "https://www.census.gov/naics/2022NAICS/2022_NAICS_Descriptions.xlsx",
    ],
    "generated_on": "2026-04-27",
    "count": len(entries),
    "levels": list(LEVELS.values()),
    "entries": entries,
}
with open(OUT_PATH, 'w') as f:
    json.dump(payload, f, indent=2, ensure_ascii=False)

print(f"\nwrote {os.path.getsize(OUT_PATH):,} bytes to {OUT_PATH}")
